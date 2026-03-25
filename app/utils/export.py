"""
Generate an Excel (.xlsx) export of attendance data.

Three sheets:
  1. Attendance Grid   — students as rows, sessions as columns, color-coded
  2. Summary           — per-student absence count, grade deduction, final score
  3. Reflections       — all reflections per student per session
"""

import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Color fills — UO Brand Colors
FILL_PRESENT   = PatternFill("solid", fgColor="D6EEDC")  # tint of Grass Green
FILL_ABSENT    = PatternFill("solid", fgColor="F4D0E1")  # tint of Berry
FILL_FLAGGED   = PatternFill("solid", fgColor="FEF9CC")  # tint of UO Yellow
FILL_HEADER    = PatternFill("solid", fgColor="104735")  # Legacy Green
FILL_WARNING   = PatternFill("solid", fgColor="FEE11A")  # UO Yellow
FILL_OVER_LIMIT= PatternFill("solid", fgColor="8D1D58")  # Berry

FONT_HEADER  = Font(color="FFFFFF", bold=True)
FONT_BOLD    = Font(bold=True)
FONT_WARNING = Font(color="8D1D58", bold=True)           # Berry for deduction warnings

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_export(course):
    """
    Generate an Excel workbook for the given Course.
    Returns a BytesIO object ready for Flask's send_file.
    """
    from app.models import Student, Session, Attendance

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    students = Student.query.filter_by(course_id=course.id).order_by(
        Student.last_name, Student.first_name
    ).all()
    sessions = Session.query.filter_by(course_id=course.id).order_by(
        Session.session_number
    ).all()

    _build_grid_sheet(wb, course, students, sessions)
    _build_summary_sheet(wb, course, students, sessions)
    _build_reflections_sheet(wb, course, students, sessions)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _build_grid_sheet(wb, course, students, sessions):
    ws = wb.create_sheet("Attendance Grid")

    # Header row
    ws.cell(1, 1, "Student").fill = FILL_HEADER
    ws.cell(1, 1).font = FONT_HEADER
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    for col_idx, session in enumerate(sessions, start=2):
        label = f"#{session.session_number}\n{session.session_date.strftime('%m/%d')}"
        cell = ws.cell(1, col_idx, label)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 9

    # Absence / deduction summary columns
    summary_col = len(sessions) + 2
    for label, col in [
        ("Absences", summary_col),
        ("Free", summary_col + 1),
        ("Excess", summary_col + 2),
        ("Deduction %", summary_col + 3),
    ]:
        cell = ws.cell(1, col, label)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 22

    # Student rows
    from app.models import Attendance
    for row_idx, student in enumerate(students, start=2):
        ws.cell(row_idx, 1, student.display_name).alignment = Alignment(horizontal="left")
        ws.cell(row_idx, 1).font = FONT_BOLD

        absence_count = 0
        for col_idx, session in enumerate(sessions, start=2):
            record = Attendance.query.filter_by(
                session_id=session.id, student_id=student.id
            ).first()

            if record and record.status == "present":
                status_char = "P"
                fill = FILL_PRESENT
                if record.flag_list:
                    fill = FILL_FLAGGED
                    status_char = "P*"
            elif record and record.status == "flagged":
                status_char = "F"
                fill = FILL_FLAGGED
            else:
                status_char = "A"
                fill = FILL_ABSENT
                absence_count += 1

            cell = ws.cell(row_idx, col_idx, status_char)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        excess = max(0, absence_count - course.free_absences)
        deduction = excess * course.deduction_per_absence

        ws.cell(row_idx, summary_col, absence_count)
        ws.cell(row_idx, summary_col + 1, course.free_absences)
        ws.cell(row_idx, summary_col + 2, excess)
        deduction_cell = ws.cell(row_idx, summary_col + 3, f"-{deduction:.1f}%")
        if excess > 0:
            deduction_cell.font = FONT_WARNING

    # Legend
    legend_row = len(students) + 3
    ws.cell(legend_row, 1, "Legend:").font = FONT_BOLD
    ws.cell(legend_row, 2, "P = Present").fill = FILL_PRESENT
    ws.cell(legend_row, 3, "A = Absent").fill = FILL_ABSENT
    ws.cell(legend_row, 4, "P* = Present (flagged)").fill = FILL_FLAGGED
    ws.cell(legend_row, 5, "F = Flagged").fill = FILL_FLAGGED


def _build_summary_sheet(wb, course, students, sessions):
    ws = wb.create_sheet("Summary")
    from app.models import Attendance

    headers = ["Student", "Total Sessions", "Present", "Absent", "Free Absences",
               "Excess Absences", "Grade Deduction", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.column_dimensions["A"].width = 25

    for row_idx, student in enumerate(students, start=2):
        total = len(sessions)
        present = Attendance.query.filter_by(
            student_id=student.id, status="present"
        ).count()
        absent = total - present
        excess = max(0, absent - course.free_absences)
        deduction = excess * course.deduction_per_absence

        notes = ""
        if excess > 0:
            notes = f"Grade reduced by {deduction:.1f}%"
        if absent > course.free_absences + 4:
            notes += " — Consider withdrawal"

        ws.cell(row_idx, 1, student.display_name).font = FONT_BOLD
        ws.cell(row_idx, 2, total)
        ws.cell(row_idx, 3, present)
        ws.cell(row_idx, 4, absent)
        ws.cell(row_idx, 5, course.free_absences)
        ws.cell(row_idx, 6, excess)
        deduction_cell = ws.cell(row_idx, 7, f"-{deduction:.1f}%" if deduction > 0 else "None")
        if deduction > 0:
            deduction_cell.font = FONT_WARNING
        ws.cell(row_idx, 8, notes)


def _build_reflections_sheet(wb, course, students, sessions):
    ws = wb.create_sheet("Reflections")
    from app.models import Attendance

    headers = ["Student", "Session #", "Date", "Reflection", "Word Count", "Flags"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.column_dimensions["D"].width = 60

    row_idx = 2
    for student in students:
        for session in sessions:
            record = Attendance.query.filter_by(
                session_id=session.id, student_id=student.id
            ).first()
            if record and record.reflection_text:
                ws.cell(row_idx, 1, student.display_name)
                ws.cell(row_idx, 2, session.session_number)
                ws.cell(row_idx, 3, session.session_date.strftime("%Y-%m-%d"))
                reflection_cell = ws.cell(row_idx, 4, record.reflection_text)
                reflection_cell.alignment = Alignment(wrap_text=True)
                ws.cell(row_idx, 5, record.word_count)
                flags = ", ".join(record.flag_list) if record.flag_list else ""
                flag_cell = ws.cell(row_idx, 6, flags)
                if flags:
                    flag_cell.fill = FILL_FLAGGED
                row_idx += 1
